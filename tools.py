"""
tools.py — standalone functions used by the LangGraph nodes.

Each tool is written so it can be unit-tested independently of the graph.
"""
from __future__ import annotations

import json
import logging
import os
import random
import re
import smtplib
import string
import time
from datetime import datetime, timezone
from email.mime.text import MIMEText
from typing import Any, Dict, List, Optional

import requests

logger = logging.getLogger("hospital_bed_agent.tools")

# --------------------------------------------------------------------------
# Config (pulled from environment — see .env.example)
# --------------------------------------------------------------------------
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
GROQ_MODEL = os.getenv("GROQ_MODEL", "llama-3.1-8b-instant")
GROQ_ENDPOINT = "https://api.groq.com/openai/v1/chat/completions"
GROQ_TIMEOUT_SECONDS = int(os.getenv("GROQ_TIMEOUT_SECONDS", "10"))

# Local Excel file used as the hospital database (replaces Google Sheets).
# Path is relative to this file's directory unless overridden in .env.
HOSPITAL_DATA_FILE = os.getenv(
    "HOSPITAL_DATA_FILE",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "Delhi_Hospital_Bed_Availability.xlsx"),
)
HOSPITAL_DATA_SHEET_NAME = os.getenv("HOSPITAL_DATA_SHEET_NAME", "Delhi Bed Availability")

SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM_ADDRESS = os.getenv("SMTP_FROM_ADDRESS", SMTP_USERNAME)

VALID_URGENCY_LEVELS = {"High", "Medium", "Low"}
VALID_BED_TYPES = {"ICU", "General", "Standard"}

# Minimal built-in fallback used only if the Excel file is missing/unreadable,
# so the agent never hard-crashes on a bad path.
_FALLBACK_HOSPITAL_ROWS: List[Dict[str, Any]] = [
    {"hospital_name": "Sample General Hospital", "zone": "Zone A",
     "available_beds": 5, "general_beds": 5, "icu_beds": 0,
     "total_beds": 50, "contact": "N/A", "last_updated": "N/A"},
]

# Simple in-memory cache for hospital search results (Part 8: 1hr cache)
_hospital_cache: Dict[str, Dict[str, Any]] = {}
_CACHE_TTL_SECONDS = 3600

# Simple in-memory rate limiter for Groq calls (Part 8: max 5/min)
_groq_call_timestamps: List[float] = []
_GROQ_RATE_LIMIT_PER_MIN = 5


# --------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def validate_patient_input(data: Dict[str, Any]) -> tuple[bool, Optional[str]]:
    """
    Validate a raw patient submission.
    Returns (is_valid, error_message). error_message is None when valid.
    """
    name = (data.get("patient_name") or "").strip()
    symptoms = (data.get("symptoms") or "").strip()
    email = (data.get("patient_email") or "").strip()
    location = (data.get("location") or "").strip()
    age = data.get("patient_age")

    if not name:
        return False, "Patient name is required."
    if not symptoms:
        return False, "Symptoms description is required."
    if not location:
        return False, "Location/Zone is required."
    if not email or not _EMAIL_RE.match(email):
        return False, "A valid email address is required."
    if age is None or not isinstance(age, (int, float)) or age <= 0:
        return False, "Age must be a positive number."

    return True, None


# --------------------------------------------------------------------------
# Groq symptom triage
# --------------------------------------------------------------------------
_TRIAGE_PROMPT_TEMPLATE = """Classify the following patient symptoms.
Return ONLY a JSON object (no markdown, no commentary) with exactly these two fields:
- "urgency_level": one of "High", "Medium", "Low"
- "bed_type_required": one of "ICU", "General", "Standard"

Symptoms: {symptoms}
"""

_DEFAULT_TRIAGE = {"urgency_level": "High", "bed_type_required": "General"}


def _check_groq_rate_limit() -> None:
    """Blocks briefly if we've exceeded 5 Groq calls in the last 60s."""
    now = time.time()
    global _groq_call_timestamps
    _groq_call_timestamps = [t for t in _groq_call_timestamps if now - t < 60]
    if len(_groq_call_timestamps) >= _GROQ_RATE_LIMIT_PER_MIN:
        oldest = _groq_call_timestamps[0]
        wait_for = 60 - (now - oldest)
        if wait_for > 0:
            logger.warning("Groq rate limit reached, sleeping %.1fs", wait_for)
            time.sleep(wait_for)
    _groq_call_timestamps.append(time.time())


def classify_symptoms_groq(symptoms: str) -> Dict[str, str]:
    """
    Call the Groq API to classify symptoms into urgency_level + bed_type_required.
    On any failure (timeout, bad JSON, missing key, non-2xx), returns safe defaults.
    """
    if not GROQ_API_KEY:
        logger.warning("GROQ_API_KEY not set; using default triage values.")
        return dict(_DEFAULT_TRIAGE)

    _check_groq_rate_limit()

    payload = {
        "model": GROQ_MODEL,
        "messages": [
            {"role": "user", "content": _TRIAGE_PROMPT_TEMPLATE.format(symptoms=symptoms)}
        ],
        "temperature": 0.3,
        "max_tokens": 150,
    }
    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
    }

    try:
        resp = requests.post(
            GROQ_ENDPOINT, headers=headers, json=payload, timeout=GROQ_TIMEOUT_SECONDS
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]

        # Model sometimes wraps JSON in markdown fences; strip if present.
        content = content.strip()
        if content.startswith("```"):
            content = re.sub(r"^```(?:json)?|```$", "", content, flags=re.MULTILINE).strip()

        result = json.loads(content)
        return {
            "urgency_level": result.get("urgency_level", _DEFAULT_TRIAGE["urgency_level"]),
            "bed_type_required": result.get("bed_type_required", _DEFAULT_TRIAGE["bed_type_required"]),
        }
    except requests.Timeout:
        logger.warning("Groq API timed out after %ss; using defaults.", GROQ_TIMEOUT_SECONDS)
        return dict(_DEFAULT_TRIAGE)
    except (requests.RequestException, KeyError, IndexError, json.JSONDecodeError) as e:
        logger.warning("Groq API call failed (%s); using defaults.", e)
        return dict(_DEFAULT_TRIAGE)


def sanitize_triage_response(raw: Dict[str, str]) -> Dict[str, str]:
    """
    Validate/normalize a raw triage dict, replacing anything invalid,
    placeholder-like, or empty with the documented defaults.
    """
    urgency = str(raw.get("urgency_level") or "").strip()
    bed_type = str(raw.get("bed_type_required") or "").strip()

    if urgency not in VALID_URGENCY_LEVELS or "urgency_level" in urgency.lower():
        urgency = _DEFAULT_TRIAGE["urgency_level"]

    if bed_type not in VALID_BED_TYPES or "bed_type_required" in bed_type.lower():
        bed_type = _DEFAULT_TRIAGE["bed_type_required"]

    return {"urgency_level": urgency, "bed_type_required": bed_type}


# --------------------------------------------------------------------------
# Hospital database (local Excel file, with in-memory fallback)
# --------------------------------------------------------------------------
_excel_rows_cache: Optional[List[Dict[str, Any]]] = None


def _load_excel_rows(force_reload: bool = False) -> List[Dict[str, Any]]:
    """
    Read hospital rows from the local Excel file (see HOSPITAL_DATA_FILE).
    Expected columns: Hospital Name, Area / Zone, Total Beds,
    General Beds Available, ICU Beds Available, Contact Number, Last Updated.
    A trailing 'TOTAL' summary row (no Area/Zone value) is skipped.
    Falls back to a minimal built-in dataset if the file is missing/unreadable.
    """
    global _excel_rows_cache
    if _excel_rows_cache is not None and not force_reload:
        return _excel_rows_cache

    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(HOSPITAL_DATA_FILE, data_only=True)
        ws = wb[HOSPITAL_DATA_SHEET_NAME] if HOSPITAL_DATA_SHEET_NAME in wb.sheetnames else wb.active

        rows: List[Dict[str, Any]] = []
        header_seen = False
        for row in ws.iter_rows(values_only=True):
            if not header_seen:
                # The real header row is the one starting with "Hospital Name"
                if row and str(row[0]).strip() == "Hospital Name":
                    header_seen = True
                continue

            name = row[0]
            zone = row[1]
            # Skip blank rows and the trailing TOTAL summary row.
            if not name or not zone:
                continue
            if str(name).strip().upper() == "TOTAL":
                continue

            total_beds = int(row[2] or 0)
            general_beds = int(row[3] or 0)
            icu_beds = int(row[4] or 0)
            contact = row[5] or "N/A"
            last_updated = row[6] or "N/A"

            rows.append({
                "hospital_name": str(name).strip(),
                "zone": str(zone).strip(),
                "total_beds": total_beds,
                "general_beds": general_beds,
                "icu_beds": icu_beds,
                # "Available beds" for the workflow's core filter = general + ICU combined,
                # matching the original spec's single Available Beds column.
                "available_beds": general_beds + icu_beds,
                "contact": str(contact),
                "last_updated": str(last_updated),
            })

        _excel_rows_cache = rows
        return rows

    except Exception as e:  # noqa: BLE001 - fall back on any read error
        logger.warning("Could not read hospital Excel file (%s); using fallback data.", e)
        _excel_rows_cache = list(_FALLBACK_HOSPITAL_ROWS)
        return _excel_rows_cache


def search_hospitals_by_zone(zone: str, use_cache: bool = True) -> List[Dict[str, Any]]:
    """
    Return all hospitals whose Area/Zone matches `zone`, ordered as found in
    the source file. Results are cached in-memory for 1 hour per zone
    (Part 8 requirement); the underlying file read is also cached until
    reload_hospital_data() is called.
    """
    now = time.time()
    if use_cache and zone in _hospital_cache:
        cached = _hospital_cache[zone]
        if now - cached["ts"] < _CACHE_TTL_SECONDS:
            return cached["rows"]

    all_rows = _load_excel_rows()
    matches = [r for r in all_rows if r.get("zone", "").strip().lower() == zone.strip().lower()]

    _hospital_cache[zone] = {"ts": now, "rows": matches}
    return matches


def list_available_zones() -> List[str]:
    """Return the distinct Area/Zone values present in the hospital data file."""
    all_rows = _load_excel_rows()
    zones = sorted({r["zone"] for r in all_rows})
    return zones


def reload_hospital_data() -> None:
    """Force the next read to re-parse the Excel file from disk (e.g. after
    it's been updated), bypassing both the row cache and the per-zone cache."""
    global _excel_rows_cache
    _excel_rows_cache = None
    _hospital_cache.clear()


def decrement_available_beds(hospital_name: str, zone: str) -> None:
    """
    Optional simulation hook: decrement the in-memory row cache's general-bed
    count for the given hospital (does not write back to the Excel file on
    disk). Kept best-effort / non-fatal.
    """
    rows = _load_excel_rows()
    for row in rows:
        if row["hospital_name"] == hospital_name and row["zone"] == zone:
            if row["general_beds"] > 0:
                row["general_beds"] -= 1
                row["available_beds"] = row["general_beds"] + row["icu_beds"]
            break
    # Invalidate the per-zone cache so subsequent searches see the update.
    _hospital_cache.pop(zone, None)


# --------------------------------------------------------------------------
# Reservation ID
# --------------------------------------------------------------------------
def generate_reservation_id() -> str:
    """Format: RES-YYYYMMDD-XXXXXX (6 random alphanumeric chars)."""
    date_part = datetime.now(timezone.utc).strftime("%Y%m%d")
    suffix = "".join(random.choices(string.digits, k=6))
    return f"RES-{date_part}-{suffix}"


# --------------------------------------------------------------------------
# Email
# --------------------------------------------------------------------------
def _send_email_smtp(to_address: str, subject: str, body: str) -> bool:
    if not SMTP_HOST or not SMTP_USERNAME or not SMTP_PASSWORD:
        logger.warning("SMTP not configured; email not sent. Would have sent:\nTo: %s\nSubject: %s\n%s",
                        to_address, subject, body)
        return False

    msg = MIMEText(body, "plain")
    msg["Subject"] = subject
    msg["From"] = SMTP_FROM_ADDRESS
    msg["To"] = to_address

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM_ADDRESS, [to_address], msg.as_string())
        return True
    except Exception as e:  # noqa: BLE001
        logger.error("Email send failed: %s", e)
        return False


def send_confirmation_email(
    to: str,
    patient_name: str,
    patient_age: int,
    patient_mobile: str,
    hospital: Dict[str, Any],
    urgency_level: str,
    bed_type: str,
    reservation_id: str,
    symptoms: str,
) -> bool:
    subject = f"Your Hospital Bed Reserved - {hospital.get('hospital_name', 'Hospital')}"
    body = f"""Hello {patient_name},

Your hospital bed has been reserved!

PATIENT INFORMATION:
Name: {patient_name}
Age: {patient_age}
Contact Number: {patient_mobile}

HOSPITAL DETAILS:
Hospital: {hospital.get('hospital_name', 'N/A')}
Zone: {hospital.get('zone', 'N/A')}
Contact: {hospital.get('contact', 'N/A')}
General Beds Available: {hospital.get('general_beds', 'N/A')}
ICU Beds Available: {hospital.get('icu_beds', 'N/A')}
Total Beds: {hospital.get('total_beds', 'N/A')}
Data Last Updated: {hospital.get('last_updated', 'N/A')}

TRIAGE CLASSIFICATION:
Urgency Level: {urgency_level}
Bed Type Required: {bed_type}

RESERVATION ID: {reservation_id}

Symptoms reported: {symptoms}

Note: This system uses a demo dataset (hospital names/locations are real,
bed counts are illustrative sample data, not live occupancy).

Thank you for using the AI-Powered Hospital Bed Allocation System.
"""
    return _send_email_smtp(to, subject, body)


def send_no_beds_notification(
    to: str,
    patient_name: str,
    zone: str,
    hospitals_in_zone: List[Dict[str, Any]],
    request_id: str,
) -> bool:
    subject = "Hospital Beds Not Currently Available - We'll Help"

    contact_lines = "\n".join(
        f"   - {h.get('hospital_name', 'Hospital')}: {h.get('contact', 'N/A')}"
        for h in hospitals_in_zone
    ) or "   - No hospitals currently listed for this zone."

    body = f"""Hello {patient_name},

Unfortunately, there are no available beds at hospitals in {zone} at this time.

RECOMMENDATIONS:
1. Check back in a few hours
2. Contact hospitals directly:
{contact_lines}

3. Try a different zone or hospital

Your symptom information has been saved. We'll prioritize your request if beds become available.

System ID: {request_id}
"""
    return _send_email_smtp(to, subject, body)
